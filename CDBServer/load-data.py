import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill
import datetime
import orm, utils

def readAllData():
    path = r'C:\Users\GaoYan\Desktop\一表同享\基层报表底数初步清单(2025.10.24).xlsx'
    wb : Workbook = openpyxl.load_workbook(path, read_only = True)
    sheet0 = wb[wb.sheetnames[0]]
    sheet1 = wb[wb.sheetnames[1]]
    rs1 = readSheet(sheet0)
    rs2 = readSheet(sheet1)
    wb.close()
    return rs1 + rs2

def readSheet(sheet):
    rs = []
    ATTRS = ['bbnc', 'fbcj', 'ssbm', 'sjx', 'sjxgs', 'tbcj', 'bsfs', 'ywxtmc', 'gxpl', 'bz', 'lxr', 'jbr', 'bm', 'ybtx_zh', 'ybtx_in', 'ybtx_mb', 'fk']
    for r, row in enumerate(sheet.rows):
        if r <= 1:
            continue
        cellA = row[2]
        if not cellA.value:
            break
        item = {} # 
        for idx, a in enumerate(ATTRS):
            cell = row[idx + 1] # skip 序号
            v = str(cell.value or '').strip()
            if v.lower() == 'excel填报': v = 'Excel填报'
            elif v.lower() == 'word填报': v = 'Word填报'
            item[a] = v
        bm : str = item['bm']
        if bm[0] >= '0' and bm[0] <= '9': bm = bm[1 : ]
        if bm[0] >= '0' and bm[0] <= '9': bm = bm[1 : ]
        item['bm'] = bm
        rs.append(item)
    return rs

def writeAllData(rs):
    for item in rs:
        orm.JcbdModel.diffSave(item)
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for item in orm.JcbdModel.select():
        item.sureTime = now
        item.save()

if __name__ == '__main__':
    rs = readAllData()
    writeAllData(rs)