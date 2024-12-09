import  openpyxl, os

sdepts = """蒲亭镇	www.dean.gov.cn/zw/03/07/ptz_184260/
宝塔乡	www.dean.gov.cn/zw/03/07/btx/
河东乡	www.dean.gov.cn/zw/03/07/hdx_184344/
丰林镇	www.dean.gov.cn/zw/03/07/flz/
高塘乡	www.dean.gov.cn/zw/03/07/gtx_184428/
林泉乡	www.dean.gov.cn/zw/03/07/lqx_184470/
聂桥镇	www.dean.gov.cn/zw/03/07/nqz_184512/
磨溪乡	www.dean.gov.cn/zw/03/07/mxx_184554/
吴山镇	www.dean.gov.cn/zw/03/07/wsz_184596/
爱民乡	www.dean.gov.cn/zw/03/07/amx_184638/
邹桥乡	www.dean.gov.cn/zw/03/07/zqx_184680/
车桥镇	www.dean.gov.cn/zw/03/07/cqz_184722/
塘山乡	www.dean.gov.cn/zw/03/07/tsx_184764/
彭山公益林场	www.dean.gov.cn/zw/03/07/pslc_184806/
向阳山生态林场	www.dean.gov.cn/zw/03/07/yyc_184848/
政府办公室	www.dean.gov.cn/zw/03/06/zfbgs_183105/
发展和改革委员会	www.dean.gov.cn/zw/03/06/fzhggwyh_183141/
教育体育局	www.dean.gov.cn/zw/03/06/jytyj_183177/
科学技术和工业信息化局	www.dean.gov.cn/zw/03/06/gyhxxhj_183249/
公安局	www.dean.gov.cn/zw/03/06/gaj_183285/
民政局	www.dean.gov.cn/zw/03/06/mzj_183321/
司法局	www.dean.gov.cn/zw/03/06/sfj_183357/
财政局	www.dean.gov.cn/zw/03/06/czj_183393/
人力资源和社会保障局	www.dean.gov.cn/zw/03/06/rlzyhshbzj_183429/
自然资源局	www.dean.gov.cn/zw/03/06/zrzyj_183465/
生态环境局	www.dean.gov.cn/zw/03/06/sthjj_183501/
住房和城乡建设局	www.dean.gov.cn/zw/03/06/zfhcxjsj_183537/
交通运输局	www.dean.gov.cn/zw/03/06/jtysj_183573/
水利局	www.dean.gov.cn/zw/03/06/slj_183609/
农业农村局	www.dean.gov.cn/zw/03/06/nyncj_183645/
商务局	www.dean.gov.cn/zw/03/06/swj_183681/
文化广电旅游局	www.dean.gov.cn/zw/03/06/whgdxwcblyj_183717/
卫生健康委员会	www.dean.gov.cn/zw/03/06/wsjkwyh_183753/
审计局	www.dean.gov.cn/zw/03/06/sjj_183789/
应急管理局	www.dean.gov.cn/zw/03/06/yjglj_183825/
退役军人事务局	www.dean.gov.cn/zw/03/06/tyjrswj_183861/
统计局	www.dean.gov.cn/zw/03/06/tjj_183897/
林业局	www.dean.gov.cn/zw/03/06/lyj_183933/
市场监督管理局	www.dean.gov.cn/zw/03/06/scjdglj_183969/
医疗保障局	www.dean.gov.cn/zw/03/06/ylbzj_184041/
城市管理局	www.dean.gov.cn/zw/03/06/csglj_184077/
行政审批局	www.dean.gov.cn/zw/03/06/xzfwzxgwh_184149/
高新技术产业园区管理委员会	www.dean.gov.cn/zw/03/06/gxjscyyqglwyh_184113/
税务局	www.dean.gov.cn/zw/03/06/swj_184185/
气象局	www.dean.gov.cn/zw/03/06/qxj_184221/
邹桥乡	www.dean.gov.cn/ztzl/zwgkzl/wszzt_246148/
吴山镇	www.dean.gov.cn/ztzl/zwgkzl/wszzt/
蒲亭镇	www.dean.gov.cn/ztzl/zwgkzl/wszzt_250622/
宝塔乡	www.dean.gov.cn/ztzl/zwgkzl/wszzt_248055/
河东乡	www.dean.gov.cn/ztzl/zwgkzl/wszzt_247840/
丰林镇	www.dean.gov.cn/ztzl/zwgkzl/wszzt_249695/
高塘乡	www.dean.gov.cn/ztzl/zwgkzl/wszzt_248270/
林泉乡	www.dean.gov.cn/ztzl/zwgkzl/wszzt_250125/
聂桥镇	www.dean.gov.cn/ztzl/zwgkzl/wszzt_245690/
磨溪乡	www.dean.gov.cn/ztzl/zwgkzl/wszzt_246119/
爱民乡	www.dean.gov.cn/ztzl/zwgkzl/wszzt_247145/
车桥镇	www.dean.gov.cn/ztzl/zwgkzl/wszzt_247625/
塘山乡	www.dean.gov.cn/ztzl/zwgkzl/wszzt_248407/
"""

depts = []
for d in sdepts.splitlines():
    if d and d.strip():
        ln = d.strip().split('\t')
        depts.append(ln)

def getDeptName(url):
    if not url:
        return ''
    for d in depts:
        if d[1] in url:
            return d[0]
    return '县级'

def findColumnHeaderIdx(sh, colName):
    for i in range(sh.max_column):
        cell = sh.cell(row = 1, column = i + 1)
        if cell.value and type(cell.value) == str and cell.value == colName:
            return i
    return -1

def run():
    print('加上责任单位；需含"错误页"列')
    path = input('输入xlsx文件地址:\n')
    if path and len(path) > 2 and path[0] == '"' and path[-1] == '"':
        path = path[1 : -1]
    wb : openpyxl.Workbook = openpyxl.load_workbook(path)
    sheet1 = wb[wb.sheetnames[0]]
    zdIdx = findColumnHeaderIdx(sheet1, '责任单位')
    if zdIdx >= 0:
        deptCol = zdIdx + 1
    else:
        deptCol = sheet1.max_column + 1
        sheet1.insert_cols(deptCol)
        sheet1.cell(1, deptCol).value = '责任单位'
    z1 = findColumnHeaderIdx(sheet1, '错误页')
    for i in range(1, sheet1.max_row):
        c = sheet1.cell(i + 1, z1 + 1)
        if not c:
            continue
        dp = getDeptName(c.value)
        cc = sheet1.cell(i + 1, deptCol)
        cc.value = dp
    pp = os.path.dirname(path)
    print(pp)
    newPath = os.path.join(pp, 'A.xlsx')
    wb.save(newPath)
    print('生成文件: ', newPath)
    wb.close()

run()