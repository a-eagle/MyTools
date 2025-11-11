import openpyxl, requests
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill
import datetime, re, json
import orm, utils

def readWorkbook():
    path = r'C:\Users\GaoYan\Desktop\一表同享\基层报表底数初步清单(2025.10.24).xlsx'
    wb : Workbook = openpyxl.load_workbook(path, read_only = True)
    sheet0 = wb[wb.sheetnames[0]]
    sheet1 = wb[wb.sheetnames[1]]
    rs1 = readSheet(sheet0)
    rs2 = readSheet(sheet1)
    wb.close()
    return rs1 + rs2

def readSheet(sheet):
    rs = []
    ATTRS = ['bbnc', 'fbcj', 'ssbm', 'sjx', 'sjxgs', 'tbcj', 'bsfs', 'ywxtmc', 'gxpl', 'gxsj', 'lxr', 'jbr', 'bm', 'ybtx_zh', 'ybtx_in', 'ybtx_mb', 'mark']
    stripNum = re.compile('\d+')
    for r, row in enumerate(sheet.rows):
        if r <= 1:
            continue
        cellA = row[2]
        if not cellA.value:
            break
        item = {} #
        for idx, a in enumerate(ATTRS):
            cell = row[idx + 1] # skip 序号
            v = str(cell.value or '').strip()
            if v.lower() == 'excel填报': v = 'Excel填报'
            elif v.lower() == 'word填报': v = 'Word填报'
            item[a] = v
        item['bm'] = stripNum.sub('', item['bm'])
        item['lxr'] = stripNum.sub('', item['lxr'])
        item['jbr'] = stripNum.sub('', item['jbr'])
        rs.append(item)
    return rs

def saveData(rs):
    for item in rs:
        orm.JcbdModel.diffSave(item)
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for item in orm.JcbdModel.select():
        item.sureTime = now
        item.save()

def writeToExcel(results):
    wb = Workbook()
    # wb.remove(wb.active)
    # ws = wb.create_sheet(sheetName)
    ws = wb.active
    for idx, title in enumerate(['', '报表名称', '发表层级', '所属部门', '数据项（字段）', '数据项个数', '填报层级', '报送方式', 
                                 '业务系统名称', '更新频率', '更新时间', '联系人', '经办人', '县直部门', '是否有一表同享账号', '是否在一表同享系统中', 
                                 '一表同享系统中模板名称', '备注']):
            c = ws.cell(row=2, column=idx+1, value=title)
            c.fill = GradientFill(stop = ('FFFF00', 'FFFF00'))
            # c.border = border
            c.font = Font(name='宋体', size = 12, bold = True)
            # c.alignment = alignCenter
    ATTRS = ('bbnc', 'fbcj', 'ssbm', 'sjx', 'sjxgs', 'tbcj', 'bsfs', 'ywxtmc', 'gxpl', 
             'gxsj', 'lxr', 'jbr', 'bm', 'ybtx_zh', 'ybtx_in', 'ybtx_mb', 'mark')
    START_ROW = 3
    for ridx, data in enumerate(results):
        for cidx, col in enumerate(ATTRS):
            ws.cell(row = ridx + START_ROW, column= 2 + cidx, value = data[col])
    wb.save(f'files/data.xlsx')

# 最新版的基层报表清单
def download():
    resp = requests.get('http://113.44.136.221:8010/api/list/JcbdModel')
    cnt = resp.content.decode()
    js = json.loads(cnt)
    results = [d for d in js if not d['isDelete']]
    writeToExcel(results)

if __name__ == '__main__':
    # rs = readWorkbook()
    # saveData(rs)
    download()