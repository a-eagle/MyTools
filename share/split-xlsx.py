# pip install openpyxl

import openpyxl, os

def buildPartFileNames(filePath, num):
    destDir = os.path.dirname(filePath)
    bn = os.path.basename(filePath)
    ext = os.path.splitext(bn)
    rs = []
    for i in range(num):
        fn = ext[0] + '-数据' + str(i + 1) + ext[1]
        fn = os.path.join(destDir, fn)
        rs.append(fn)
    #print('[buildPartFileNames] rs=', rs)
    return rs

def buildSheetFileNames(filePath):
    destDir = os.path.dirname(filePath)
    bn = os.path.basename(filePath)
    ext = os.path.splitext(bn)
    wb = openpyxl.load_workbook(filePath)
    rs = [os.path.join(destDir, ext[0] + sh.title  + '-数据' + ext[1]) for sh in wb.worksheets]
    wb.close()
    print('[buildSheetFileNames] rs=', rs)
    return rs

def delRows(sheet, saveRows, titlesRowNum):
    print(saveRows)
    sheet.delete_rows(saveRows[1] + titlesRowNum + 2, sheet.max_row)
    if saveRows[0] > 0:
        sheet.delete_rows(titlesRowNum + 1, saveRows[0] + titlesRowNum - 1)

def splitExcelFileByData(filePath, titlesRowNum, splitFileNum):
    filePath = os.path.abspath(filePath)
    print(filePath)
    pfn = buildPartFileNames(filePath, splitFileNum)
    print(pfn)
    wb = openpyxl.load_workbook(filePath)
    sheet = wb.worksheets[0]
    avgNum = (sheet.max_row - titlesRowNum) // splitFileNum
    print('avgNum=', avgNum)
    wb.close()
    print('max row data: ', sheet.max_row)
    for i in range(splitFileNum):
        wb = openpyxl.load_workbook(filePath)
        sheet = wb.worksheets[0]
        if sheet.max_row <= titlesRowNum + splitFileNum:
            raise Exception('[splitExcelFileData] 表格行数太少')
        if i + 1 >= splitFileNum:
            saveRows = (avgNum * i, sheet.max_row)
        else:
            saveRows = (avgNum * i, avgNum * (i + 1) - 1)
        delRows(sheet, saveRows, titlesRowNum)
        wb.save(pfn[i])
        wb.close()

def delSheets(wb, saveSheetIdx):
    for i in range(len(wb.worksheets) - 1, -1, -1):
        if i != saveSheetIdx:
            wb.remove(wb.worksheets[i])

def splitExcelFileBySheet(filePath):
    filePath = os.path.abspath(filePath)
    pfn = buildSheetFileNames(filePath)
    for i, pf in enumerate(pfn):
        wb = openpyxl.load_workbook(filePath)
        delSheets(wb, i)
        wb.save(pf)
        wb.close()

if __name__ == '__main__':
    fp = r'C:\Users\GaoYan\Desktop\2023\共享数据\0. 行政许可\12交运局\1\2023年交运局行政执法月报.xlsx'
    #splitExcelFileByData(fp, 1, 3)
    splitExcelFileBySheet(fp)