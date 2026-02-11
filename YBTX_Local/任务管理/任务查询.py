import peewee as pw, json, requests, datetime, time, urllib.parse, os, sys, re

import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill

sys.path.append(__file__[0 : __file__.upper().index('任务管理')])
import decrypt, login
from orm import *
from utils import *

class ProgressManager:
    def __init__(self) -> None:
        self.tasks = []

    def filter(self, *filters):
        # filter
        num = 0
        for q in TaskModel.select().order_by(TaskModel.createTime.desc()):
            js = json.loads(q.cnt)
            cnd = True
            for f in filters:
                cnd = cnd and f(js)
            if cnd:
                self.tasks.append(q)
                num += 1
                # print(f'[filter] accept [{num :2d}]', 'C' + q.createTime[0 : 10], 'E' + q.deadlineTime[0 : 10], q.statusDesc, q.title)

    def parse(self):
        progress = {}
        for t in self.tasks:
            tp = TaskProgress(t)
            tp.parseTask(progress)
        slistXZ = []
        slistDM = []
        for k in progress:
            if '乡' in k or '镇' in k or '场' in k:
                slistXZ.append(progress[k])
            else:
                slistDM.append(progress[k])
        slistXZ.sort(key = lambda x: x['deptName'])
        slistDM.sort(key = lambda x: x['deptName'])
        results = slistDM + slistXZ
        return results

    def getNum(self, obj, k):
        if (k not in obj) or (not obj[k]):
            return 0
        return obj[k]
    
    def calcCommentWidth(self, row):
        w = 0
        for ch in row:
            if ord(ch) < 255:
                w += 8
            else:
                w += 14
        return w
    
    def addComment(self, cell, obj, k):
        if k not in obj:
            return
        rows = []
        for idx, info in enumerate(obj[k]):
            task = info['task']
            users = info.get('users', None)
            row = f'【{idx + 1}】任务创建时间：{task.createTime[0 : 10]}  任务名称：{task.title} '
            if users:
                row += '未填报人员：(' + '、'.join([u['nickname'] for u in users]) + ')   '
            rows.append(row)
        rowsWidth = [self.calcCommentWidth(row) for row in rows]
        txt = ' | '.join(rows)
        width = max(rowsWidth)
        if width > 650: width = 650
        cell.comment = Comment(txt, None, height = 17 * len(obj[k]), width = width)

    def getNum_s(self, obj, ks):
        num = 0
        for k in ks:
            v = self.getNum(obj, k)
            if v: num += 1
        return num

    def log(self, all = False):
        f = open('任务统计/rs.log', 'w')
        print('\t单位\t填报人员\t待处理\t已转发\t待审核', file = f)
        no = 0
        for k in self.results:
            if (not all) and (not self.getNum_s(k, ['待处理', '已转发', '待审核', '已审核'])):
                continue
            if '测试' in k['nickName']:
                continue
            no += 1
            print(no, k['deptName'], k['nickName'], self.getNum(k, '待处理'),
                  self.getNum(k, '已转发'), self.getNum(k, '待审核'), sep = '\t', file = f)
        f.close()

    def writeExcel_乡镇填报情况(self, wb : Workbook, results, startTime, endTime, sheetName, mainTitle):
        results = [r for r in results if r['deptName'][-1] in ('乡', '镇', '场')]
        results.sort(key = lambda x : x['deptName'].encode('gbk'))
        day = datetime.date.today().strftime('%Y-%m-%d')
        ws = wb.create_sheet(sheetName)
        side = Side(border_style='thin', color='000000')
        bfont = Font(name='宋体', size=12)
        bfont2 = Font(name='宋体', size=12, color='ff0000', bold=True)
        bfont3 = Font(name='宋体', size=12, color='00B050', bold=True)
        border = Border(left=side, right=side, top=side, bottom=side)
        alignCenter = Alignment(horizontal='center', vertical='center')
        ws.merge_cells("A1:I1")
        a1 = ws['A1']
        a1.value = mainTitle
        a1.font = Font(name='宋体', size=14, bold=True)
        a1.alignment = alignCenter
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 33
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12

        ws.merge_cells("A2:I2")
        a2 = ws['A2']
        a2.value = f'(统计范围: 任务结束时间在{startTime}至{endTime}之间)'
        a2.font = bfont2
        a2.alignment = alignCenter

        for idx, title in enumerate(['', '单位', '填报人员', '待处理', '已填报', '待审核', '已审核', '村(社区)\n未填报', '村(社区)\n已填报']):
            c = ws.cell(row=3, column=idx+1, value=title)
            c.fill = GradientFill(stop = ('FFFF00', 'FFFF00'))
            c.border = border
            c.font = Font(name='宋体', size = 12, bold = True)
            c.alignment = alignCenter
        no = 0
        START_ROW = 3
        for idx, k in enumerate(results):
            # if (not all) and (not self.getNum_s(k, ['待处理', '已转发', '待审核', '已审核'])):
            #     continue
            if '测试' in k['nickName']:
                continue
            # check is empty line
            allNum = 0
            for x in ('待处理', '已处理', '待审核', '已审核', '下级待处理', '下级已处理'):
                allNum += self.getNum(k, x)
            if allNum == 0:
                continue
            no += 1
            c1 = ws.cell(row=no + START_ROW, column=1, value=no)
            c2 = ws.cell(row=no + START_ROW, column=2, value=k['deptName'])
            c3 = ws.cell(row=no + START_ROW, column=3, value=k['nickName'])
            c4 = ws.cell(row=no + START_ROW, column=4, value=self.getNum(k, '待处理'))
            c5 = ws.cell(row=no + START_ROW, column=5, value=self.getNum(k, '已处理'))
            c6 = ws.cell(row=no + START_ROW, column=6, value=self.getNum(k, '待审核'))
            c7 = ws.cell(row=no + START_ROW, column=7, value=self.getNum(k, '已审核'))
            c8 = ws.cell(row=no + START_ROW, column=8, value=self.getNum(k, '下级待处理'))
            c9 = ws.cell(row=no + START_ROW, column=9, value=self.getNum(k, '下级已处理'))
            self.addComment(c4, k, '待处理-明细')
            self.addComment(c6, k, '待审核-明细')
            self.addComment(c8, k, '下级待处理-明细')
            for cx in (c1, c2, c3, c4, c5, c6, c7, c8, c9):
                cx.font = bfont
                cx.border = border
            for cx in (c1, c3, c4, c5, c6, c7, c8, c9):
                cx.alignment = alignCenter
            for cx in (c4, c6, c8):
                if cx.value != 0:
                    cx.font = bfont2
            for cx in (c5, c7, c9):
                if cx.value != 0:
                    cx.font = bfont3

    def print_乡镇填报人员汇总(self, results):
        results = [r for r in results if r['deptName'][-1] in ('乡', '镇', '场')]
        rs = {}
        for d in results:
            dept = d['deptName']
            user = d['nickName']
            if dept not in rs: rs[dept] = {'未完成': set(), '已完成':set()}
            if d.get('待处理', 0) or d.get('待审核', 0) or d.get('下级待处理', 0):
                rs[dept]['未完成'].add(user)
            else:
                rs[dept]['已完成'].add(user)
        depts = list(rs.keys())
        depts.sort(key = lambda x : x.encode('gbk'))
        for k in depts:
            deptName = k.replace('德安县', '')
            deptName = f'【{deptName}】'
            print(deptName, '  填报已完成(', '、'.join(rs[k]['已完成']),')', ';  填报未完成(', '、'.join(rs[k]['未完成']),')', sep='')

    def print_部门审核情况(self):
        progress = {}
        for t in self.tasks:
            deptName = t.deptName + '-' + t.nickName # 任务下发单位 + 姓名
            if deptName not in progress:
                progress[deptName] = {}
            tp = TaskProgress(t)
            for n in tp.nodes:
                if n['pnodeId']: continue
                if n['statusDesc'] != '待审核': continue
                key = str(n['taskId']) + '/' + t.title
                if key not in progress[deptName]:
                    progress[deptName][key] = []
                dname = n['deptName']
                if '（' in dname: dname = dname[0 : dname.index('（')]
                progress[deptName][key].append(dname)
        print('-------------------部门待审核数-----------------------------')
        for p in progress:
            info = progress[p]
            if not info:
                continue
            print(p, '-->', progress[p])
            num = 0
            for k in info:
                num += len(info[k])
            print(p.replace('德安县委', '').replace('德安县', ''), ': ', '待审核数', num)
        pass

    def print_任务完成率(self, tag):
        if not self.tasks:
            print(f'{tag}任务总数: 0')
            return
        num = 0
        for t in self.tasks:
            if t.statusDesc == '已完成':
                num += 1
        print(f'{tag}任务总数：{len(self.tasks)} 已完成：{num} 完成率：{int(num / len(self.tasks) * 100)}%')

class TaskMgr:
    def __init__(self) -> None:
        self.tasks = []

    def filter(self, *filters):
        # filter
        self.tasks = []
        for q in TaskModel.select():
            js = json.loads(q.cnt)
            cnd = True
            for f in filters:
                cnd = cnd and f(js)
            if cnd:
                self.tasks.append(q)

    # 统计任务下发情况
    def print_部门任务汇总(self):
        rs = {}
        for r in self.tasks:
            if r.deptName[-1] in '乡镇场':
                continue
            if r.deptName not in rs:
                rs[r.deptName] = []
            #print(r.deptName, json.loads(r.cnt)['createTime'], r.title)
            rs[r.deptName].append(r.title)
        print('----------部门任务下发情况--------------------')
        print(' ' * 30, '月量')
        total = 0
        for dept in rs:
            item = rs[dept]
            total += len(item)
            dept = dept + ' ' * ((12 - len(dept)) * 2)
            if item:
                print(dept, len(item), sep = '\t')
            else:
                print(dept, 0, sep = '\t')
        print('总数：', total)

    def print_非临时任务(self):
        self.tasks.sort(key = lambda t : t.createTime, reverse= True)
        fls, ls, nf = [], [], []
        for task in self.tasks:
            cnt = json.loads(task.cnt)
            taskTypeDesc = cnt.get('taskTypeDesc', '')
            if taskTypeDesc == '高频报表任务':
                fls.append((task.createTime, taskTypeDesc, task.title))
                continue
            if taskTypeDesc != '自定义任务':
                continue
            tmps = json.loads(task.refTemplate)
            tmp = tmps[0]
            find = LocalTemplateModel.get_or_none(name = tmp['templateTitle'])
            if find:
                if find.peroid != '临时性':
                    fls.append((task.createTime, taskTypeDesc, find.peroid, task.title))
                else:
                    ls.append((task.createTime, taskTypeDesc, find.peroid, task.title, '|||', tmp['templateTitle'], task.deptName))
            else:
                nf.append((task.createTime, taskTypeDesc, task.title, '///', tmp['templateTitle']))
        print('【非临时任务数】', len(fls))
        print('【临时任务数】', len(ls))
        for r in ls:
            print('\t', r)
        print('【未知任务数】', len(nf))
        for r in nf:
            print('\t', r)

def print_村社区填报任务量():
    cun = {}
    for task in TaskModel.select().order_by(TaskModel.createTime.desc()):
        if not task.progress:
            continue
        ps = json.loads(task.progress)
        for p in ps:
            deptName = p['deptName']
            sname : str = deptName[deptName.index('（') + 1 : deptName.index('）')]
            if sname.startswith('德安县>'):
                continue
            sname = sname.replace('德安县', '')
            item = cun.get(sname, None)
            if not item: item = cun[sname] = {'任务总数' : 0, '填报数量': 0}
            item['任务总数'] += 1
            if p['firstSubmitTime']:
                item['填报数量'] += 1
    cunList = []
    for c in cun:
        tt = {'name': c}
        tt.update(cun[c])
        cunList.append(tt)
    cunList.sort(key= lambda k : k['任务总数'], reverse= True)
    for i, c in enumerate(cunList):
        print(c)
    import 任务管理.dept as dept
    cunNames = []
    for c in dept.Cun.select():
        cunNames.append(c.xzName + '>' + c.cunName)
    for c in cunList:
        if c['name'] in cunNames and c['填报数量']:
            cunNames.remove(c['name'])
    print('未填报村：', cunNames)
    cunList.sort(key= lambda k : k['name'].encode('gbk'), reverse= False)
    pass

def print_村社区使用模板最多():
    cun = {} # key = templateTitle , val: num
    for task in TaskModel.select().order_by(TaskModel.createTime.desc()):
        if not task.progress or not task.refTemplate:
            continue
        temp = json.loads(task.refTemplate)
        if not temp:
            continue
        temp = temp[0]
        ps = json.loads(task.progress)
        num = 0
        for p in ps:
            deptName = p['deptName']
            sname : str = deptName[deptName.index('（') + 1 : deptName.index('）')]
            if sname.startswith('德安县>'):
                continue
            num += 1
        cun[temp['templateTitle']] = cun.get(temp['templateTitle'], 0) + num
    cs = [(k, cun[k]) for k in cun]
    cs.sort(key = lambda c : c[1], reverse = True)
    for i in range(0, 10):
        print(cs[i])
    pass

def fmtDate(day):
    if not day:
        return None
    if isinstance(day, datetime.date):
        day = day.strftime('%Y-%m-%d')
    day = day.strip()
    if len(day) == 8:
        return day[0 : 4] + '-' + day[4 : 6] + '-' + day[6 : 8]
    if len(day) == 10:
        return day
    return None

def print_乡镇统计时间段(startTime, endTime, soonEndTime = None):
    startTime = fmtDate(startTime)
    endTime = fmtDate(endTime)
    soonEndTime = fmtDate(soonEndTime)
    if not startTime or not endTime or startTime > endTime:
        print('开始日期/结束日期错误')
        return
    TODAY = datetime.date.today().strftime('%Y-%m-%d')
    proMgr = ProgressManager()
    proMgr.filter(
             lambda x: x['deadlineTime'][0 : 10] >= startTime,
             lambda x: x['deadlineTime'][0 : 10] <= endTime,
             lambda x: x['statusDesc'] in ('填报中', '已完成'),
             )
    results = proMgr.parse()
    wb = Workbook()
    wb.remove(wb.active)
    proMgr.writeExcel_乡镇填报情况(wb, results, startTime, endTime, '全部', f'乡镇"一表同享"填报任务完成情况({TODAY})')
    proMgr.print_乡镇填报人员汇总(results)
    unResults = [r for r in results if r.get('待处理', 0) or r.get('待审核', 0) or r.get('下级待处理', 0) ]
    proMgr.writeExcel_乡镇填报情况(wb, unResults, startTime, endTime, '未完成', '已超期任务填报情况')
    proMgr.print_任务完成率(f'乡镇{startTime}至{endTime} ')

    if soonEndTime and soonEndTime > endTime:
        proMgr = ProgressManager()
        et = datetime.datetime.strptime(endTime, '%Y-%m-%d') + datetime.timedelta(days = 1)
        ets = et.strftime('%Y-%m-%d')
        proMgr.filter(
             lambda x: x['deadlineTime'][0 : 10] >= ets,
             lambda x: x['deadlineTime'][0 : 10] <= soonEndTime,
             lambda x: x['statusDesc'] in ('填报中', '已完成'),
             )
        results = proMgr.parse()
        proMgr.writeExcel_乡镇填报情况(wb, results, ets, soonEndTime, '即将到期', '即将到期任务填报情况')
        proMgr = ProgressManager()
        proMgr.filter(
             lambda x: x['deadlineTime'][0 : 10] >= startTime,
             lambda x: x['deadlineTime'][0 : 10] <= soonEndTime,
             lambda x: x['statusDesc'] in ('填报中', '已完成'),
             )
        proMgr.print_任务完成率(f'乡镇{startTime}至{soonEndTime} ')
    wb.save(f'files/乡镇填报任务完成情况({TODAY}).xlsx')

def print_县直部门待审核任务(startTime, endTime):
    proMgr = ProgressManager()
    proMgr.filter(
             lambda x: x['deadlineTime'][0 : 10] >= startTime,
             lambda x: x['deadlineTime'][0 : 10] <= endTime,
             lambda x: x['statusDesc'] in ('填报中'),
             )
    proMgr.tasks.sort(key = lambda task: task.deptName + '-' + task.nickName)
    no = 1
    for task in proMgr.tasks:
        tp = TaskProgress(task)
        if tp.checkFullFinished():
            print(f'[{no :2d}]', '待审核已完成任务->', task.deptName, task.nickName, 'C' + task.createTime[0 : 10], 'E' + task.deadlineTime[0 : 10], task.title)
            no += 1
    print('------------------------------------------------------')

def 部门使用数量(startDay = '2025-01-01', endDay = datetime.date.today().strftime('%Y-%m-%d')):
    xzCun = set()
    bm = set()
    createTaskDepts = set()
    from 任务管理.dept import Cun
    for cun in Cun.select():
        xzCun.add(cun.xzName)
        xzCun.add(cun.cunName)
    xzCun.add('向阳山生态林场')
    for task in TaskModel.select().where(TaskModel.createTime >= startDay, TaskModel.createTime <= endDay, TaskModel.statusDesc != '已终止'): # 已终止
        dp = task.deptName.replace('德安县', '')
        if dp and (dp not in xzCun):
            # is bu meng
            bm.add(task.deptName)
            createTaskDepts.add(task.deptName)
        if not task.progress:
            continue
        nodes = json.loads(task.progress)
        for node in nodes:
            if node['statusDesc'] != '已审核':
                continue
            dp = node['deptName']
            dpf = dp[0 : dp.index('（')]
            dp = dpf.replace('德安县', '')
            if dp and (dp not in xzCun):
                bm.add(dpf)
    bm = list(bm)
    bm.sort(key = lambda k: k.encode('gbk'))
    createTaskDepts = list(createTaskDepts)
    createTaskDepts.sort(key = lambda k: k.encode('gbk'))
    return bm, createTaskDepts

def print_市周报表数据(year, month, weekStart, weekEnd):
    TODAY = datetime.date.today().strftime('%Y-%m-%d')
    month = f'{int(month) :02d}'
    yearTaskNum = TaskModel.select(pw.fn.count()).where(TaskModel.createTime >= str(year), TaskModel.statusDesc != '已终止').scalar()
    monthTaskNum = TaskModel.select(pw.fn.count()).where(TaskModel.createTime >= f'{year}-{month}', TaskModel.statusDesc != '已终止').scalar()
    weekTaskNum = TaskModel.select(pw.fn.count()).where(TaskModel.createTime >= weekStart, TaskModel.createTime <= weekEnd, TaskModel.statusDesc != '已终止').scalar()
    print('-------------------市周报数据----------------------------')
    print('年度任务分发数量: ', yearTaskNum)
    print('本月新增任务数量: ', monthTaskNum)
    print('本周新增任务数量: ', weekTaskNum)
    dm = DeptDownloader()
    # dm.loadNetDatas()
    dm.loadDatas()
    cbm, tbm = dm.部门使用数量()
    # print('----年度部门使用数量（全部）------------\n', f'【{len(bm)}】', '、'.join(bm))
    print('----年度部门使用数量（创建任务）-------- \n', f'【{len(cbm)}】', '、'.join(cbm))
    print('----年度部门使用数量（仅填报）---------- \n', f'【{len(tbm)}】', '、'.join(tbm))

    #taskMgr.print_非临时任务()

class DeptTaskManager:
    def __init__(self) -> None:
        self.jcbbData = None
        self.results = None

    def loadServerJcbb(self):
        def strToHex(s : str):
            HEX = '0123456789ABCDEF'
            bs = s.encode()
            vals = []
            for b in bs:
                b = int(b)
                l, h = b & 0xf, (b >> 4) & 0xf
                vals.append(HEX[h])
                vals.append(HEX[l])
            return ''.join(vals)
        filters = strToHex(json.dumps([
            {'col': 'fbcj', 'op': '==', 'val': '县区级'}, {'col': 'isDelete', 'op': '==', 'val': '0'},
        ]))
        resp = requests.get(f'http://113.44.136.221:8010/api/list/JcbdModel?filters={filters}')
        js = resp.json()
        rs = {}
        for it in js:
            dept = it['bm']
            if rs.get(dept, None) == None:
                rs[dept] = []
            rs[dept].append(it)
        self.jcbbData = rs
        
    def calcSum(self, year, month):
        self.results = {}
        for dept in self.jcbbData:
            self.results[dept] = {
                        '部门': dept,
                        '全部模板': self.jcbbData[dept], '当月模板': [],
                        '已发任务': [], '已发任务模板': [],
                        '更新频率': self.get_更新频率(self.jcbbData[dept])}
        self.calc_当月模板(month)
        self.calc_已发任务_已发任务模板(year, month)

    def get_更新频率(self, temps):
        SS = ['年报', '半年报', '季报', '月报', '阶段性', '临时性', '实时更新']
        rs = {}
        for s in SS:
            for t in temps:
                if t['gxpl'] == s:
                    rs[s] = rs.get(s, 0) + 1
        for t in temps:
            if t['gxpl'] not in SS:
                rs['未知'] = rs.get('未知', 0) + 1
        return rs

    def calc_当月模板(self, month):
        TAG = f'{month}月'
        def inMonth(gxsj):
            if TAG not in gxsj:
                return False
            if TAG == '1月' or TAG == '2月':
                gxsj = gxsj.replace('11月', ' ')
                gxsj = gxsj.replace('12月', ' ')
                return TAG in gxsj
            return True

        for dept in self.results:
            obj = self.results[dept]
            for tp in obj['全部模板']:
                if (tp['gxpl'] == '月报') or inMonth(tp['gxsj']):
                    obj['当月模板'].append(tp)

    def matchTemplate(self, fullTemplate, simpleTemplate : str):
        rep = simpleTemplate.replace('+', '.*?')
        rs = re.findall(rep, fullTemplate)
        return len(rs) > 0

    def calc_已发任务模板(self, info, task):
        if not task.refTemplate:
            return
        taskTemp = json.loads(task.refTemplate)
        for tp in taskTemp:
            info['已发任务模板'].append(tp)
            tp['REF-TASK'] = task
            refTitle = tp['templateTitle']
            # print('[calcTemplateTask] ', info['部门'], '任务:', task.title, '==> 模板:', refTitle)
            isInCurMonth = False
            for ctp in info['当月模板']:
                match = self.matchTemplate(refTitle, ctp['ybtx_mb'])
                if match:
                    ctp['是否下发任务'] = True
                    ctp['REF-TASK'] = task
                    isInCurMonth = True
            tp['是否是当月模板'] = isInCurMonth
            # find in 
            for it in info['全部模板']:
                match = self.matchTemplate(refTitle, it['ybtx_mb'])
                if match:
                    tp['REF_SYS_MODEL'] = it
                    continue

    def calc_已发任务_已发任务模板(self, year, month):
        taskMgr = TaskMgr()
        taskMgr.filter(
                lambda x: x['createTime'] >= f'{year}-{month :02d}-01',
                lambda x: x['createTime'] <= f'{year}-{month :02d}-31',
                lambda x: x['statusDesc'] in ('填报中', '已完成', '已下发，未开始'),
                )
        for task in taskMgr.tasks:
            dn = task.deptName
            if dn[-1] in '乡镇场':
                continue
            dept = self.simpleDeptName(dn)
            if dept not in self.results:
                self.results[dept] = {'部门': dept, '全部模板': [], '当月模板': [], '已发任务': [], '已发任务模板': [], '更新频率': {}}
            info = self.results[dept]
            info['已发任务'].append(task)
            self.calc_已发任务模板(info, task)

    # 应发任务数
    def getTemplateNum(self, item):
        if not item:
            return 0
        obj = item.get('当月模板', None)
        if not obj:
            return 0
        return len(obj)

    # 实发任务数
    def getRealTemplateNum(self, item):
        if not item:
            return 0
        obj = item.get('已发任务模板', None)
        if not obj:
            return 0
        return len(obj)

    def writeExcel(self, month):
        wb = Workbook()
        ws = wb.active
        self._writeSumInfo(ws, month)
        ws2 = wb.create_sheet('当月模板')
        self._writeCurMonth(ws2, month)
        wb.save(f'files/任务下发统计表_{datetime.date.today()}.xlsx')

    def _writeCurMonth(self, ws, month):
        side = Side(border_style='thin', color='000000')
        border = Border(left=side, right=side, top=side, bottom=side)
        alignCenter1 = Alignment(horizontal='left', vertical='center', wrapText = True)
        alignCenter2 = Alignment(horizontal='center', vertical='center', wrapText = True)
        row = 1
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 20
        for idx, title in enumerate(['序号', '责任单位', '表单名称', '更新频率', '更新时间', '是否下发任务', '关联任务', '关联任务时间']):
            c = ws.cell(row = row, column = idx + 1, value = title)
            c.fill = GradientFill(stop = ('acb9ca', 'acb9ca'))
            c.border = border
            c.font = Font(name='宋体', size = 11, bold = True)
            c.alignment = alignCenter2
        
        for dept in self.results:
            obj = self.results[dept]
            temps = obj.get('当月模板', None)
            if not temps:
                continue
            for tp in temps:
                row += 1
                ws.row_dimensions[row].height = 25
                haveSendTask = '是' if tp.get('是否下发任务', False) else '否'
                refTask = tp.get('REF-TASK', '')
                taskTitle = ''
                ctime = ''
                if refTask:
                    taskTitle = refTask.title
                    ctime = refTask.createTime[0 : 10]
                vals = [row - 1, dept, tp['bbnc'], tp['gxpl'], tp['gxsj'], haveSendTask, taskTitle, ctime]
                for idx, val in enumerate(vals):
                    c = ws.cell(row = row, column = idx + 1, value = val)
                    c.border = border
                    if idx == len(vals) - 3 and haveSendTask == '否':
                        c.font = Font(name='宋体', size = 11, bold = False, color = 'ff00ff')
                    else:
                        c.font = Font(name='宋体', size = 11, bold = False)
                    c.alignment = alignCenter1

        # 非当月任务模板
        ws.row_dimensions[row].height = 25
        ws.merge_cells(f"A{row}:H{row}")
        cell = ws.cell(row = row, column = 1, value = '非当月任务模板')
        cell.fill = GradientFill(stop = ('acb9ca', 'acb9ca'))
        for dept in self.results:
            obj = self.results[dept]
            temps = obj.get('已发任务模板', None)
            if not temps:
                continue
            for tp in temps:
                if tp['是否是当月模板']:
                    continue
                row += 1
                ws.row_dimensions[row].height = 25
                refTask = tp.get('REF-TASK', '')
                refSysModel = tp.get('REF_SYS_MODEL', None)
                gxpl, gxsj = '', ''
                if refSysModel:
                    gxpl = refSysModel['gxpl']
                    gxsj = refSysModel['gxsj']
                vals = [row - 2, dept, tp['templateTitle'], gxpl, gxsj, '', refTask.title, refTask.createTime[0 : 10]]
                for idx, val in enumerate(vals):
                    c = ws.cell(row = row, column = idx + 1, value = val)
                    c.border = border
                    c.font = Font(name='宋体', size = 11, bold = False)
                    c.alignment = alignCenter1

    def _writeSumInfo(self, ws, month):
        side = Side(border_style='thin', color='000000')
        bfont = Font(name='宋体', size=11)
        bfont2 = Font(name='宋体', size=11, color='ff0000', bold=True)
        bfont3 = Font(name='宋体', size=11, color='00B050', bold=True)
        bfont4 = Font(name='黑体', size=11)
        border = Border(left=side, right=side, top=side, bottom=side)
        alignCenter = Alignment(horizontal='center', vertical='center')
        alignCenter2 = Alignment(horizontal='center', vertical='center', wrapText = True)
        alignCenter3 = Alignment(horizontal='left', vertical='center')
        ws.merge_cells("A1:I1")
        a1 = ws['A1']
        today = datetime.date.today()
        a1.value = f'德安县“一表同享”业务表单、部门任务下发统计（{today.year}年{today.month}月{today.day}日）'
        a1.font = Font(name='方正小标宋简体', size=16, bold=False)
        a1.alignment = alignCenter
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 30
        ws.row_dimensions[4].height = 70
        ws.row_dimensions[5].height = 30
        ws.row_dimensions[6].height = 30
        ws.row_dimensions[7].height = 30

        def COL(idx): return chr(ord('A') + idx)

        ws.column_dimensions['A'].width = 15
        for i in range(len(self.results)):
            ws.column_dimensions[COL(i + 1)].width = 10

        ws.merge_cells(f"A2:{COL(len(self.results))}2")
        a2 = ws[f'A2']
        a2.value = f'（县区级）'
        a2.font = Font(name='方正小标宋简体', size=15, color='000000', bold=False)
        a2.fill = GradientFill(stop = ('acb9ca', 'acb9ca'))
        a2.alignment = alignCenter3
        a2.border = border
        for am in ws[f'A2:{COL(len(self.results))}2'][0]:
            am.border = border

        row = 3
        for idx, title in enumerate(['', '表单总数', f'{month}月应发\n表单数', f'{month}月实发\n表单数', f'{month}月\n完成率']):
            sf = ws.cell(row=row, column=1, value=title)
            sf.fill = GradientFill(stop = ('acb9ca', 'acb9ca'))
            sf.border = border
            sf.font = Font(name='宋体', size = 11, bold = True)
            sf.alignment = alignCenter2
            row += 1

        def GXPL(s):
            rr = ''
            for k in s:
                k2 = k
                if k == '实时更新': k2 = '实时'
                rr += k2 +  str(s[k]) + '\n'
            return rr.strip()
        
        def finishRate(m, c):
            if m == 0:
                return '' if c == 0 else '100%'
            if c >= m:
                return '100%'
            return f'{int(c / m * 100)}%'

        for idx, dept in enumerate(self.results):
            row = 3
            cur = self.results[dept]
            yf = self.getTemplateNum(cur)
            sf = self.getRealTemplateNum(cur)
            rate = finishRate(yf, sf)
            if yf == 0: yf = ''
            if not sf and not yf:
                sf = ''
            vals = [dept, GXPL(cur['更新频率']), yf, sf, rate]
            for v in vals:
                sf = ws.cell(row=row, column=idx+2, value=v)
                sf.border = border
                if row == 3:
                    sf.font = Font(name='黑体', size = 11, bold = False)
                    sf.fill = GradientFill(stop = ('acb9ca', 'acb9ca'))
                else:
                    sf.font = Font(name='宋体', size = 11, bold = True)
                sf.alignment = alignCenter2
                row += 1

    def simpleDeptName(self, n):
        if '人力资源和社会保障局' in n: return '县人社局'
        if '住房与城乡建设局' in n: return '县住建局'
        if '县卫生健康委员会' in n: return '县卫健委'
        return n.replace('德安', '')

def 核对村社区使用数量():
    dm = DeptDownloader()
    dm.loadDatas()

    cuns = dm.get_村社区()
    cunNames = [n['deptName'] for n in cuns]
    print('系统上的村社区：', len(cunNames), cunNames, '\n')
    for c in cuns:
        k = c['deptName']
        if k[-3 : ] == '村委会' or k[-3 : ] =='居委会':
            continue
        print(k, '==>', c['ancestorsName'])

    xzcInfos = {}
    for c in cuns:
        cn = c['deptName']
        xzName = dm.getTopDept(cn)
        xzcInfos[xzName] = xzcInfos.get(xzName, {})
        xzcInfos[xzName][cn] = []
    for it in RecvTaskModel.select().where(RecvTaskModel.isDelete == False).order_by(RecvTaskModel.firstSubmitTime.desc()): #.where(RecvTaskModel.createTime >= '2025-12-01')
        if it.deptName not in cunNames:
            continue
        xzName = dm.getTopDept(it.deptName)
        xzcInfos[xzName][it.deptName].append(it)

    print('-----tasks num---------- ')
    idx = 1
    for xz in xzcInfos:
        for cc in xzcInfos[xz]:
            tasks = xzcInfos[xz][cc]
            print(idx, xz, cc, len(tasks), "" + tasks[0].firstSubmitTime[0 : 10], tasks[0].title, sep='    ')
            idx += 1
    pass

def 核对是否有空部门_空村社区():
    userDepts = {}
    depts = []
    for it in UserModel.select():
        dd = json.loads(it.depts)
        for m in dd:
            dn = m['deptName']
            userDepts[dn] = userDepts.get(dn, 0) + 1
    dm = DeptDownloader()
    dm.loadDatas()
    for b in dm.datas:
        deptName = dm.datas[b]['deptName']
        depts.append(deptName)
    for idx, d in enumerate(depts):
        num = userDepts.get(d, 0)
        if num == 0:
            print('=====空部门==========>', d)
        else:
            print(f'[{idx + 1}]', d, '==>', num)

def addDay(day, daysNum):
    return day + datetime.timedelta(days = daysNum)

def main():
    核对村社区使用数量()


    TODAY = datetime.date.today()
    #print_村社区填报任务量()
    #print_村社区使用模板最多()

    # 下载任务、进度
    if True:
        # authorization, decryptKey = window.key4
        downloader = TaskDownloader()
        downloader.enableUpdate = 1
        downloader.loadTasks()
        downloader.loadTemplate()
        downloader.loadProgress()
        # startTime = input('开始日期:')
        # endTime = input('结束日期:') or TODAY
        # soonEndTime = input('即将到期日期:') or addDay(TODAY, 2)
        # print_乡镇统计时间段(startTime = '2025-09-01', endTime = addDay(TODAY, 0), soonEndTime = addDay(TODAY, 5))
        # print_县直部门待审核任务(startTime = '2025-06-01', endTime = '2099-12-31')

        YEAR, MONTH = 2026, 2
        mgr = DeptTaskManager()
        mgr.loadServerJcbb()
        mgr.calcSum(YEAR, MONTH)
        mgr.writeExcel(MONTH)
        pass

    #-----------------------------------------------------
    if True:
        print('---------------TaskRecvDownloader---------------------------')
        recv = TaskRecvDownloader(enableUpdate = True)
        recv.loadTasksFromRecv()
        year = TODAY.year
        month = TODAY.month
        weekStart = fmtDate(TODAY - datetime.timedelta(TODAY.weekday()))
        weekEnd = fmtDate(TODAY)
        print_市周报表数据(year, month, weekStart, weekEnd)

if __name__ == '__main__':
    # 核对村社区使用数量()
    main()