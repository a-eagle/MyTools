import peewee as pw, requests, json, sys, os, time

import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill

sys.path.append(os.path.dirname(os.path.dirname(__file__)))

import decrypt, login

db = pw.SqliteDatabase('台账/房屋.db')

class HouseModel(pw.Model):
    s_id = pw.CharField(null = True, default = '') # 601204009100842
    createTime = pw.CharField(null = True, default = '') # 2024-12-30
    updateTime = pw.CharField(null = True, default = '') # 2025-02-11
    createBy = pw.CharField(null = True, default = '') # DeAn003
    updateBy = pw.CharField(null = True, default = '') # DeAn003
    delFlag = pw.CharField(null = True, default = '') # 0
    dsRegionCode = pw.CharField(null = True, default = '') # 360426
    regionCode = pw.CharField(null = True, default = '') # 360426200101
    uniqueId = pw.CharField(null = True, default = '') # H1873634531420815457
    houseAddress = pw.CharField(null = True, default = '') # 杨桥至共青路南侧（杨桥新城二区）11幢2单元602号
    houseOwnerName = pw.CharField(null = True, default = '') # 钱立鸿
    houseOwnerPhone = pw.CharField(null = True, default = '') # 18779270209
    houseOwnerId = pw.CharField(null = True, default = '') # 36**************35
    residenceType = pw.CharField(null = True, default = '') # 1
    unitNumber = pw.CharField(null = True, default = '') # 360426001044GB00001F00110046
    rightsNumber = pw.CharField(null = True, default = '') # 赣(2020)德安县不动产权第0000287号
    rightsCertificateNumber = pw.CharField(null = True, default = '') # 36005485231
    registrationInstitution = pw.CharField(null = True, default = '') # 德安县不动产登记局
    houseRightsType = pw.CharField(null = True, default = '') # 国有建设用地使用权/房屋所有权
    houseFloorArea = pw.CharField(null = True, default = '') # 124.19
    houseNature = pw.CharField(null = True, default = '') # 99
    regionCodeLabel = pw.CharField(null = True, default = '') # 江西省>九江市>德安县>宝塔乡>向阳社区
    regionCodeProvince = pw.CharField(null = True, default = '') # 江西省
    regionCodeCity = pw.CharField(null = True, default = '') # 九江市
    regionCodeTown = pw.CharField(null = True, default = '') # 德安县
    regionCodeCommunity = pw.CharField(null = True, default = '') # 宝塔乡
    regionCodeVillage = pw.CharField(null = True, default = '') # 向阳社区
    regionCodeGrid = pw.CharField(null = True, default = '') # 

    class Meta:
        database = db

class BuildingModel(pw.Model):
    s_id= pw.CharField(null = True, default = '') #  580301661171664
    createTime= pw.CharField(null = True, default = '') #  
    updateTime= pw.CharField(null = True, default = '') #  
    createBy= pw.CharField(null = True, default = '') #  DeAn003
    updateBy= pw.CharField(null = True, default = '') #  DeAn003
    delFlag= pw.CharField(null = True, default = '') #  0
    dsRegionCode= pw.CharField(null = True, default = '') #  360426
    regionCode= pw.CharField(null = True, default = '') #  360426104200
    uniqueId= pw.CharField(null = True, default = '') #  B1851459684720959499
    name= pw.CharField(null = True, default = '') #  360426002008JC00696F00010001
    buildingAddress= pw.CharField(null = True, default = '') #  吴山镇何铺村二组
    buildingOwnerName= pw.CharField(null = True, default = '') #  周建霞
    houseUsage= pw.CharField(null = True, default = '') #  1
    buildingArea= pw.CharField(null = True, default = '') #  215.877平方米
    regionCodeLabel= pw.CharField(null = True, default = '') #  江西省>九江市>德安县>吴山镇>何铺村委会
    regionCodeProvince= pw.CharField(null = True, default = '') #  江西省
    regionCodeCity= pw.CharField(null = True, default = '') #  九江市
    regionCodeTown= pw.CharField(null = True, default = '') #  德安县
    regionCodeCommunity= pw.CharField(null = True, default = '') #  吴山镇
    regionCodeVillage= pw.CharField(null = True, default = '') #  何铺村委会
    regionCodeGrid= pw.CharField(null = True, default = '') #  
    class Meta:
        database = db

db.create_tables([HouseModel, BuildingModel])

class DownloadMgr:
    def __init__(self) -> None:
        self.authorization = None

    def login(self):
        self.authorization = login.login(decrypt.DECRYPT_KEY)

    # pageIdx = [1, ...]
    def download(self, pageIdx, type, model):
        PAGE_SIZE = 100
        headers = {'authorization': self.authorization, 'accept': "application/json, text/plain, */*"}
        url = f'http://10.8.52.17:8088/ledger-be/ledger/{type}/pageList'
        params = {"current": pageIdx, "size": PAGE_SIZE, "latestUpdateType":0, "tagIds": []}
        resp = requests.post(url, json = params, headers = headers)
        txt = resp.content.decode('utf-8')
        js = json.loads(txt)
        if js['code'] != 200:
            print('Download Error', js)
            return None
        data = decrypt.decrypt(js['data'])
        js = json.loads(data)
        pages = (js['total'] + PAGE_SIZE - 1) // PAGE_SIZE
        datas = js['records']
        rs = []
        for d in datas:
            em = model(**d)
            em.s_id = d['id']
            rs.append(em)
        model.bulk_create(rs, batch_size = 100)
        return pages
    
    # type = building | house
    def downloadAll(self, type, model):
        pageNum = 10
        curPage = 1
        while curPage <= pageNum:
            print('[download]', curPage)
            pageNum = self.download(curPage, type, model)
            curPage += 1
            time.sleep(5)

def readExcel():
    path = r'D:\vscode\MyTools\YBTX_Local\台账\T_360499_DAXFW.xlsx'
    wb : Workbook = openpyxl.load_workbook(path, read_only = True)
    ns = wb.sheetnames[0]
    sheet = wb[ns]
    rows = {}
    for idx, row in enumerate(sheet.rows):
        if idx <= 2:
            continue
        qylx = str(row[1].value).strip()
        if qylx and len(qylx) > 1:
            continue
        item = {}
        item[1] = '德安县'
        item[4] = row[0].value.strip() # 房屋详细地址
        item[10] = row[2].value.strip() #  不动产单元号
        item[11] = str(row[3].value).strip() # 不动产权证字号
        item[13] = row[4].value.strip() # 登记机构
        item[16] = str(row[6].value).strip() # 房屋建筑面积
        rows[item[10]] = item
    wb.close()
    return rows

def compare(datas):
    ex = {}
    for it in HouseModel.select():
        ex[it.unitNumber] = it
        if it.unitNumber in datas:
            del datas[it.unitNumber]
    for it in BuildingModel.select():
        ex[it.name] = it
        if it.name in datas:
            del datas[it.name]
    print('[compare] insert num = ', len(datas))

def write(data):
    DEST = r'C:\Users\GaoYan\Desktop\房屋台账导入模板.xlsx'
    wb = openpyxl.open(DEST)
    sheet = wb[wb.sheetnames[0]]
    no = 2
    for k in data:
        no += 1
        row = data[k]
        for c in row:
            sheet.cell(row = no, column = c, value = row[c])
    wb.save(DEST)
    wb.close()


if __name__ == '__main__':
    a = 232896 + 61485 + 25715 + 92737 + 28449
    b = 4364 + 88 + 1610 + 1056 + 539 + 661 + 2135 + 423 + 823
    print(a, a + b)
    mgr = DownloadMgr()
    # mgr.login()
    # mgr.downloadAll('house', HouseModel)
    # mgr.downloadAll('building', BuildingModel)
    datas = readExcel()
    compare(datas)
    write(datas)
