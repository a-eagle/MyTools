import peewee as pw, requests, json, sys, os, time

import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill

sys.path.append(os.path.dirname(os.path.dirname(__file__)))

import decrypt, login

db = pw.SqliteDatabase('台账/企业.db')

class EnterpriseModel(pw.Model):
    s_id = pw.CharField(null = True, default = '') # 546588537713228
    createTime = pw.CharField(null = True, default = '')# 2024-07-25 17:09:14
    updateTime = pw.CharField(null = True, default = '')#   2025-12-09 12:54:51
    createBy = pw.CharField(null = True, default = '') # DeAn002
    updateBy = pw.CharField(null = True, default = '') #  DeAn003
    delFlag = pw.CharField(null = True, default = '') #0
    dsRegionCode = pw.CharField(null = True, default = '')#360426
    regionCode = pw.CharField(null = True, default = '')#360426
    uniqueId = pw.CharField(null = True, default = '') # E1816400304854241280
    name = pw.CharField(null = True, default = '')  # 九江市互动投资管理中心（有限合伙）
    businessCreditCode = pw.CharField(null = True, default = '')  # 91360426343215310T
    legalPerson = pw.CharField(null = True, default = '')  #  古予舟
    enterpriseRegistrationNumber = pw.CharField(null = True, default = '')  #  360426310004760
    registrationAuthority = pw.CharField(null = True, default = '')  #  德安县市场监督管理局
    registeredCapital = pw.CharField(null = True, default = '')  #  300万
    registeredAddress = pw.CharField(null = True, default = '')  #  江西省九江市德安县吴山镇农机厂
    industry = pw.CharField(null = True, default = '')  #  11
    businessStatus = pw.CharField(null = True, default = '')  #  2
    mainBusinessScope = pw.CharField(null = True, default = '')  #  设计、制作、发布、代理国内外各类广告；投资咨询（不含期货、金融、证券咨询）；创业管理服务（依法须经批准的项目，经相关部门批准后方可开展经营活动）**
    isListed = pw.CharField(null = True, default = '')  #  # 0
    enterpriseMigrationTime = pw.CharField(null = True, default = '')  #  #  2015-05-25
    l1 = pw.CharField(null = True, default = '')  #  九江市宇威投资管理中心（委派代表：古予舟）
    regionCodeLabel = pw.CharField(null = True, default = '')  #   江西省>九江市>德安县
    regionCodeProvince = pw.CharField(null = True, default = '')  #  
    regionCodeCity = pw.CharField(null = True, default = '')  #  
    regionCodeTown = pw.CharField(null = True, default = '')  #  
    regionCodeCommunity = pw.CharField(null = True, default = '')  #  
    regionCodeVillage = pw.CharField(null = True, default = '')  #  
    regionCodeGrid = pw.CharField(null = True, default = '')  #  
    class Meta:
        database = db

db.create_tables([EnterpriseModel])

class EnterpriseMgr:
    def __init__(self) -> None:
        self.authorization = None

    def login(self):
        self.authorization = login.login(decrypt.DECRYPT_KEY)
    
    # pageIdx = [1, ...]
    def download(self, pageIdx):
        PAGE_SIZE = 100
        headers = {'authorization': self.authorization, 'accept': "application/json, text/plain, */*"}
        url = f'http://10.8.52.17:8088/ledger-be/ledger/enterprise/pageList'
        params = {"current": pageIdx, "size": PAGE_SIZE, "latestUpdateType":0, "tagIds": []}
        resp = requests.post(url, json = params, headers = headers)
        txt = resp.content.decode('utf-8')
        js = json.loads(txt)
        if js['code'] != 200:
            print('Download Error', js)
            return None
        data = decrypt.decrypt(js['data'])
        js = json.loads(data)
        pages = (js['total'] + PAGE_SIZE - 1) // PAGE_SIZE
        datas = js['records']
        rs = []
        for d in datas:
            em = EnterpriseModel(**d)
            em.s_id = d['id']
            rs.append(em)
        EnterpriseModel.bulk_create(rs, batch_size = 100)
        return pages
    
    def downloadAll(self):
        pageNum = 10
        curPage = 2
        while curPage <= pageNum:
            print('[download]', curPage)
            pageNum = self.download(curPage)
            curPage += 1
            time.sleep(5)

def getXZ(addr):
    pass

def readExcel():
    path = r'D:\vscode\MyTools\YBTX_Local\台账\DAX_FR.xlsx'
    wb : Workbook = openpyxl.load_workbook(path, read_only = True)
    ns = wb.sheetnames[0]
    sheet = wb[ns]
    rows = []
    for idx, row in enumerate(sheet.rows):
        if idx <= 1:
            continue
        item = {}
        item[0] = '德安县'
        item[3] = row[3].value.strip() # 统一社会信用代码
        item[4] = row[1].value.strip() # 企业名称
        item[5] = row[4].value.strip() # 法人/联系人
        item[7] = row[2].value.strip() # 企业注册号
        addr = item[10] = row[12].value.strip() # 企业注册地址
        item[19] = row[10].value.strip() # 主营业务范围
        item[1] = ''
        item[2] = ''
        rows.append(item)
    wb.close()
    return rows

if __name__ == '__main__':
    # mgr = EnterpriseMgr()
    # mgr.login()
    # mgr.downloadAll()
    pass
    ext = {}
    for d in EnterpriseModel.select():
        ext[d.businessCreditCode] = d
    ds = readExcel()
    datas = []
    DEST = r'C:\Users\GaoYan\Desktop\企业台账导入模板.xlsx'
    wb = openpyxl.open(DEST)
    sheet = wb[wb.sheetnames[0]]
    no = 1
    for idx, d in enumerate(ds):
        if d[3] in ext:
            continue
        datas.append(d)
        no += 1
        for c in d:
            sheet.cell(row = no, column = c + 1, value = d[c])
    wb.save(DEST)
    wb.close()
    # 69218 条